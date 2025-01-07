import streamlit as st
import pandas as pd
import numpy as np
import yfinance as yf
from datetime import datetime
from dateutil.relativedelta import relativedelta
import time
import datetime as dt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl import load_workbook


def main():
    st.title("Momentum Streamlit App")

    def getMedianVolume(data):
        return (round(data.median(), 0))

    def getDailyReturns(data):
        return (data.pct_change(fill_method='ffill'))  # Modified

    # change inf value to max and min value of that column
    def getMaskDailyChange(data):
        m1 = getDailyReturns(data).eq(np.inf)  # for inf value
        m2 = getDailyReturns(data).eq(-np.inf)  # for -inf value
        return (getDailyReturns(data).mask(m1, df[~m1].max(), axis=1).mask(m2, df[~m2].min(), axis=1).bfill(axis=1))

    def getStdev(data):
        return (np.std(getMaskDailyChange(data) * 100))

    def getStdRatio(data, data1):
        return ((getStdev(data) / getStdev(data1) * 100))  # to judge volatility 1 month against 1 year

    def getAbsReturns(data):
        x = (data.iloc[-1] / data.iloc[0] - 1) * 100
        return (round(x, 2))

    def getVolatility(data):
        return (round(np.std(data) * np.sqrt(252) * 100, 2))

    def getMonthlyPrices(data):
        grps = data.groupby([data.index.year, data.index.month])
        monthlyPrices = pd.DataFrame()
        for k in grps:
            monthlyPrices = pd.concat([monthlyPrices, k[1].tail(1)])
        # monthlyPrices = monthlyPrices.append(k[1].tail(1))
        return monthlyPrices

    def getMonthlyReturns(data):
        return (data.pct_change())

    def getSharpe(data):
        return (round(np.sqrt(252) * data.mean() / data.std(), 2))

    def getSortino(data):
        return (np.sqrt(252) * data.mean() / data[data < 0].std())

    def getMaxDrawdown(data):
        cummRet = (data + 1).cumprod()
        peak = cummRet.expanding(min_periods=1).max()
        drawdown = (cummRet / peak) - 1
        return drawdown.min()

    def getCalmar(data):
        return (data.mean() * 252 / abs(getMaxDrawdown(data)))

    def getAbsMomentumVolAdjusted(absReturn, volatility):
        return (absReturn / volatility)

    def getNMonthRoC(data, N):
        ret = round((data.iloc[-1] / data.iloc[-1 - N] - 1) * 100, 2)
        return (ret)

    def getNWeekRoC(data, N):
        ret = round((data.iloc[-1] / data.iloc[-1 - N] - 1) * 100, 2)
        return (ret)

    def getFIP(data):
        retPos = np.sum(data.pct_change()[1:] > 0)
        retNeg = np.sum(data.pct_change()[1:] < 0)
        return (retPos - retNeg)

    def getSharpeRoC(roc, volatility):
        return (round(roc / volatility, 2))

    # Beta should be calculated against relevant Index instead of Nifty50?
    def getBeta(dfNifty, data12M):

        dailyReturns = getDailyReturns(pd.concat([dfNifty, data12M], axis=1))[1:]

        var = np.var(dailyReturns['Nifty'])  # Modified

        cov = dailyReturns.cov()

        cols = cov.columns[1:]

        beta = []

        for k in cols:
            beta.append(round(cov.loc[k, 'Nifty'] / var, 2))

        return beta

    # Streamlit Layout
    st.title("Momentum Ranking App")

    # To suppress future warnings about using ffill method in pct_change()
    import warnings
    warnings.simplefilter(action='ignore', category=FutureWarning)

    # Dropdown options with display labels and corresponding values
    ranking_options = {
        "sharpe12M": "sharpe12M",
        "sharpe3M": "sharpe3M",
        "avgSharpe 12M/6M/3M": "avgSharpe",
        "avgSharpe 12M/9M/6M/3M": "avg_All"
    }
    # Display dropdown for ranking method selection
    ranking_method_display = st.selectbox(
        "Select Ranking Method",
        options=list(ranking_options.keys()),  # Display labels
        index=0  # Default to the first option
    )

    # Get the actual value for the selected display label
    ranking_method = ranking_options[ranking_method_display]

    # Select Universe with default value as 'N750'
    universe = ['Nifty50', 'Nifty100', 'Nifty200', 'Nifty250', 'Nifty500', 'N750', 'AllNSE']
    U = st.selectbox('Select Universe:', universe, index=5)  # Default value is 'N750' (index 5)

    # Date Picker for Lookback Start Date
    selected_date = st.date_input("Select Lookback Starting Date", datetime.today())
    dt2 = datetime.strptime(str(selected_date), "%Y-%m-%d").strftime('%Y-%m-%d')

    # Displaying Date Range Information
    dates = {
        'startDate': datetime.strptime('2000-01-01', '%Y-%m-%d'),
        'endDate': datetime.strptime(dt2, '%Y-%m-%d'),
        'date12M': datetime.strptime(dt2, '%Y-%m-%d') - relativedelta(months=12),
        'date9M': datetime.strptime(dt2, '%Y-%m-%d') - relativedelta(months=9),
        'date6M': datetime.strptime(dt2, '%Y-%m-%d') - relativedelta(months=6),
        'date3M': datetime.strptime(dt2, '%Y-%m-%d') - relativedelta(months=3),
        'date1M': datetime.strptime(dt2, '%Y-%m-%d') - relativedelta(months=1),
    }

    st.write("##### Date Range:")
    st.write(f"Start Date: **{dates['startDate'].strftime('%d-%m-%Y')}**")
    st.write(f"End Date: **{dates['endDate'].strftime('%d-%m-%Y')}**")

    # Read index file based on selected universe
    if U == 'N750':
        file_path = 'https://raw.githubusercontent.com/prayan2702/Streamlit-momn/refs/heads/main/ind_niftytotalmarket_list.csv'
    elif U == 'AllNSE':
        file_path = f'https://raw.githubusercontent.com/prayan2702/Streamlit-momn/refs/heads/main/NSE_EQ_ALL.csv'
    else:
        file_path = f'https://raw.githubusercontent.com/prayan2702/Streamlit-momn/refs/heads/main/ind_{U.lower()}list.csv'

    df = pd.read_csv(file_path)
    df['Yahoo_Symbol'] = df.Symbol + '.NS'
    df = df.set_index('Yahoo_Symbol')
    symbol = list(df.index)

    # Add a button to start the process
    start_button = st.button("Start Data Download")

    if start_button:
        # Download data when the button is pressed
        CHUNK = 50
        close = []
        high = []
        volume = []

        # Create a progress bar
        progress_bar = st.progress(0)
        status_text = st.empty()  # Placeholder for progress text

        # Track the number of stocks downloaded
        total_symbols = len(symbol)
        chunk_count = (total_symbols // CHUNK) + (1 if total_symbols % CHUNK != 0 else 0)

        # Use a separate list to handle failed downloads
        failed_symbols = []

        for k in range(0, len(symbol), CHUNK):
            _symlist = symbol[k:k + CHUNK]

            # Try downloading data for each chunk of symbols
            try:
                _x = yf.download(_symlist, start=dates['startDate'], progress=False)
                close.append(_x['Close'])
                high.append(_x['High'])
                volume.append(_x['Close'] * _x['Volume'])
            except Exception as e:
                failed_symbols.extend(_symlist)  # Add failed symbols to the list
                st.write(f"Failed to download data for: {', '.join(_symlist)}. Error: {e}")

            # Update progress bar after each chunk
            progress = (k + CHUNK) / total_symbols
            progress = min(max(progress, 0.0), 1.0)  # newly added for progress bar error
            progress_bar.progress(progress)

            # Update status text with progress percentage
            progress_percentage = int(progress * 100)
            status_text.text(f"Downloading... {progress_percentage}%")

            time.sleep(0.5)

        # After the download is complete, update the progress bar and text
        progress_bar.progress(1.0)
        status_text.text("Download complete!")

        # **********************************
        # Function to calculate next rebalance date
        def get_next_rebalance_date(current_date):
            current_date = pd.Timestamp(current_date)

            # List of months for rebalance: March and September
            rebalance_months = [3, 9]

            rebalance_dates = []
            for month in rebalance_months:
                # Get last day of the month
                last_day = pd.Timestamp(current_date.year, month, 1) + pd.offsets.MonthEnd(0)
                # Adjust if it's a weekend
                while last_day.weekday() >= 5:  # Saturday (5) or Sunday (6)
                    last_day -= dt.timedelta(days=1)
                rebalance_dates.append(last_day)

            # Check for next rebalance date
            for date in sorted(rebalance_dates):
                if date > current_date:
                    return date.date()  # Return as datetime.date

            # If no date is valid in current year, calculate for next year
            next_year_rebalance = []
            for month in rebalance_months:
                last_day = pd.Timestamp(current_date.year + 1, month, 1) + pd.offsets.MonthEnd(0)
                while last_day.weekday() >= 5:
                    last_day -= dt.timedelta(days=1)
                next_year_rebalance.append(last_day)

            return sorted(next_year_rebalance)[0].date()  # Return as datetime.date

        # Test date
        current_date = dt.date.today()
        next_rebalance_date = get_next_rebalance_date(current_date)

        # Format the date as DD-MM-YYYY
        formatted_rebalance_date = next_rebalance_date.strftime("%d-%m-%Y")

        # List of applied filters
        filters = [
            "Volume greater than 1 crore (volm_cr > 1)",
            "Closing price above 200-day moving average (Close > dma200d)",
            "12-month rate of change (ROC) greater than 6.5% (roc12M > 6.5)",
            "Number of circuit hits in a year less than 20 (circuit < 20)",
            "Stock is within 25% of its all-time high (AWAY_ATH > -25)",
            "12-month return less than 1000% (roc12M < 1000)",
            "1-month ROC to 12-month ROC ratio less than 50% (roc1M / roc12M * 100 < 50)",
            "Closing price above ₹30 (Close > 30)",
            "No more than 10 circuits of 5% in the last 3 months (circuit5 <= 10)",
        ]

        # Sidebar section
        with st.sidebar:
            st.header("Menu")

            # Display the rebalance date
            st.info(f"Index Rebalance Date: **{formatted_rebalance_date}**.")
            with st.expander("Applied Filters", expanded=False):
                st.write("The following conditions are applied:")
                for i, filter_desc in enumerate(filters, start=1):
                    st.write(f"{i}. {filter_desc}")

        # **************************************************
        # Handle failed symbols (if any)
        if failed_symbols:
            st.write(f"Failed to download data for the following symbols: {', '.join(failed_symbols)}")

        # Convert close, high, and volume lists to DataFrames
        close = pd.concat(close, axis=1) if close else pd.DataFrame()
        high = pd.concat(high, axis=1) if high else pd.DataFrame()
        volume = pd.concat(volume, axis=1) if volume else pd.DataFrame()

        # Ensure the index is datetime
        close.index = pd.to_datetime(close.index)
        high.index = pd.to_datetime(high.index)
        volume.index = pd.to_datetime(volume.index)

        data20Y = close.loc[:dates['endDate']].copy()
        volume20Y = volume.loc[:dates['endDate']].copy()
        high20Y = high.loc[:dates['endDate']].copy()
        volume12M = volume20Y.loc[dates['date12M']:].copy()

        # At least 12 months of trading is required
        data12M = data20Y.loc[dates['date12M']:].copy()
        data9M = data20Y.loc[dates['date9M']:].copy()
        data6M = data20Y.loc[dates['date6M']:].copy()
        data3M = data20Y.loc[dates['date3M']:].copy()
        data1M = data20Y.loc[dates['date1M']:].copy()

        # ******************

        # Calculate metrics for dfStats (e.g., 1 month momentum, 3-month volatility, etc.)
        dfStats = pd.DataFrame(index=symbol)
        dfStats['Close'] = round(data12M.iloc[-1], 2)
        data12M_Temp = data12M.fillna(0)
        dfStats['dma200d'] = round(data12M_Temp.rolling(window=200).mean().iloc[-1], 2)  # 200-day DMA

        # Rate of change

        dfStats['roc12M'] = getAbsReturns(data12M)
        dfStats['roc9M'] = getAbsReturns(data9M)
        dfStats['roc6M'] = getAbsReturns(data6M)
        dfStats['roc3M'] = getAbsReturns(data3M)
        dfStats['roc1M'] = getAbsReturns(data1M)

        # Volatility

        dfStats['vol12M'] = getVolatility(getDailyReturns(data12M))
        dfStats['vol9M'] = getVolatility(getDailyReturns(data9M))
        dfStats['vol6M'] = getVolatility(getDailyReturns(data6M))
        dfStats['vol3M'] = getVolatility(getDailyReturns(data3M))

        dfStats['sharpe12M'] = getSharpeRoC(dfStats['roc12M'], dfStats['vol12M'])
        dfStats['sharpe9M'] = getSharpeRoC(dfStats['roc9M'], dfStats['vol9M'])
        dfStats['sharpe6M'] = getSharpeRoC(dfStats['roc6M'], dfStats['vol6M'])
        dfStats['sharpe3M'] = getSharpeRoC(dfStats['roc3M'], dfStats['vol3M'])

        # dfStats['avgSharpe'] = (dfStats[["sharpe12M", "sharpe6M", "sharpe3M"]].mean(axis=1)).round(2)  # 1st Factor
        # ****************************************
        # Columns for different ranking methods
        columns_avgSharpe = ["sharpe12M", "sharpe6M", "sharpe3M"]
        columns_avgAll = ["sharpe12M", "sharpe9M", "sharpe6M", "sharpe3M"]

        # Conditional logic based on ranking_method
        if ranking_method == "avgSharpe":
            dfStats['avgSharpe'] = dfStats[columns_avgSharpe].mean(axis=1).round(2)
        elif ranking_method == "avg_All":
            dfStats['avg_All'] = dfStats[columns_avgAll].mean(axis=1).round(2)
        # ******************************************

        dfStats['volm_cr'] = (getMedianVolume(volume12M) / 1e7).round(2)

        # ***************************

        # Calculate ATH and Away from ATH% (Additional added condition)
        dfStats['ATH'] = round(high20Y.max(), 2)
        dfStats['AWAY_ATH'] = round((dfStats['Close'] / dfStats['ATH'] - 1) * 100,
                                    2)  # Calculate %away from ALL TIME HIGH

        # Calculate No. of circuits in one year (Additional added condition)
        dataDaily_pct = round(getDailyReturns(data12M) * 100, 2)
        dfStats['circuit'] = (dataDaily_pct == 4.99).sum() + (dataDaily_pct == 5.00).sum() + (
                dataDaily_pct == 9.99).sum() + (dataDaily_pct == 10.00).sum() + (dataDaily_pct == 19.99).sum() + (
                                     dataDaily_pct == 20.00).sum() + (dataDaily_pct == -4.99).sum() + (
                                     dataDaily_pct == -5.00).sum() + (dataDaily_pct == -9.99).sum() + (
                                     dataDaily_pct == -10.00).sum() + (dataDaily_pct == -19.99).sum() + (
                                     dataDaily_pct == -20.00).sum()

        # Calculate No. of 5% circuit in last 3-month for filter
        dataDaily_pct5 = round(getDailyReturns(data3M) * 100, 2)
        dfStats['circuit5'] = (dataDaily_pct5 == 4.99).sum() + (dataDaily_pct5 == 5.00).sum() + (
                dataDaily_pct5 == -4.99).sum() + (dataDaily_pct5 == -5.00).sum()

        # Add 'Ticker' as a reset index column and rename it
        dfStats = dfStats.reset_index().rename(columns={'index': 'Ticker'})

        # Convert to String and Remove the .NS Suffix
        dfStats['Ticker'] = dfStats['Ticker'].astype(str)
        dfStats['Ticker'] = dfStats['Ticker'].str.replace('.NS', '', case=False, regex=False)

        # Handle Nan and inf values to zero(0) for ranking
        if ranking_method == "avgSharpe":
            dfStats['avgSharpe'] = dfStats['avgSharpe'].replace([np.inf, -np.inf], np.nan).fillna(0)
        elif ranking_method == "avg_All":
            dfStats['avg_All'] = dfStats['avg_All'].replace([np.inf, -np.inf], np.nan).fillna(0)
        dfStats['sharpe12M'] = dfStats['sharpe12M'].replace([np.inf, -np.inf], np.nan).fillna(0)
        dfStats['sharpe3M'] = dfStats['sharpe3M'].replace([np.inf, -np.inf], np.nan).fillna(0)

        # Add Rank column based on 'avgSharpe' and sort by Rank
        # dfStats['Rank'] = dfStats[ranking_method].rank(ascending=False,method='first').astype(int)
        # dfStats = dfStats.sort_values('Rank').set_index('Rank')  # Set 'Rank' as index
        # *******************************************************
        # Sort by primary(ranking_method) and secondary columns(roc12M) (if no-ties then rank by primary column, otherwise use primary and then secondary to break tie)
        # Conditional logic for sorting
        if ranking_method in ["avg_All", "sharpe12M"]:
            dfStats = dfStats.sort_values(by=[ranking_method, 'roc12M'], ascending=[False, False])
        elif ranking_method in ["avgSharpe", "sharpe3M"]:
            dfStats = dfStats.sort_values(by=[ranking_method, 'roc3M'], ascending=[False, False])

        # Assign unique ranks based on the sorted order
        dfStats['Rank'] = range(1, len(dfStats) + 1)

        # Set 'Rank' as the index
        dfStats = dfStats.set_index('Rank')
        # **********************************************************

        # Show both filtered and unfiltered data in Streamlit
        st.write("Unfiltered Data with Rank:")
        st.write(dfStats)

        # Apply conditions
        cond1 = dfStats['volm_cr'] > 1  # volume filter greater than 1 crore
        cond3 = dfStats['Close'] > dfStats['dma200d']  # above 200-day DMA
        cond4 = dfStats['roc12M'] > 6.5  # 12 month Rate of Change > 6.5%
        cond5 = dfStats['circuit'] < 20  # No. of circuit less than 20 (Additional added condition)
        cond6 = dfStats['AWAY_ATH'] > -25  # Away from All Time High within 25%
        cond7 = dfStats['roc12M'] < 1000  # 12M return less than 10x and atleast 12M trading required
        cond8 = dfStats['roc1M'] / dfStats['roc12M'] * 100 < 50  # 1m ROC/12m ROC < 50%
        cond9 = dfStats['Close'] > 30  # stock pricce above 30
        cond10 = dfStats['circuit5'] <= 10  # No. of 5% circuit in last 3 month should be less than 11 (newly added)

        # Create final momentum filter column
        dfStats['final_momentum'] = cond1 & cond3 & cond4 & cond5 & cond6 & cond7 & cond8 & cond9 & cond10

        # *************************************

        # Filter stocks meeting all conditions
        filtered = dfStats[dfStats['final_momentum']].sort_values('Rank', ascending=True)

        st.write("Filtered Data with Rank:")
        st.write(filtered)

        # ***********************************************************
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        def format_excel(file_name):
            # Open the written file using openpyxl
            wb = openpyxl.load_workbook(file_name)
            ws = wb.active

            # Add Borders to All Cells
            thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                                 bottom=Side(style="thin"))

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Freeze the top row
            ws.freeze_panes = 'A2'

            # Define the bold font style
            bold_font = Font(bold=True)

            # Format headers
            header_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue
            header_font = Font(bold=True, color="FFFFFF")
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Automatically adjust column widths based on content
            for col in ws.columns:
                max_length = 0
                column = col[0].column
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2
                ws.column_dimensions[cell.column_letter].width = adjusted_width

            # Define cell color for cells that do not meet filter conditions
            no_condition_fill = PatternFill(start_color="d6b4fc", end_color="d6b4fc", fill_type="solid")
            bold_font = Font(bold=True)

            # Get the headers and find column indexes by name
            headers = [cell.value for cell in ws[1]]
            col_indices = {
                'volm_cr': headers.index('volm_cr') + 1,
                'Close': headers.index('Close') + 1,
                'dma200d': headers.index('dma200d') + 1,
                'AWAY_ATH': headers.index('AWAY_ATH') + 1,
                'roc12M': headers.index('roc12M') + 1,
                'circuit': headers.index('circuit') + 1,
                'roc1M': headers.index('roc1M') + 1,
                'circuit5': headers.index('circuit5') + 1,
                'Ticker': headers.index('Ticker') + 1,
                'Rank': headers.index('Rank') + 1
            }

            # Apply conditional formatting
            for row in range(2, ws.max_row + 1):
                condition_failed = False
                if (volume := ws.cell(row=row, column=col_indices['volm_cr']).value) is not None and volume < 1:
                    ws.cell(row=row, column=col_indices['volm_cr']).fill = no_condition_fill
                    ws.cell(row=row, column=col_indices['volm_cr']).font = bold_font
                    condition_failed = True
                if (close := ws.cell(row=row, column=col_indices['Close']).value) is not None and close <= ws.cell(
                        row=row,
                        column=
                        col_indices[
                            'dma200d']).value:
                    ws.cell(row=row, column=col_indices['Close']).fill = no_condition_fill
                    ws.cell(row=row, column=col_indices['Close']).font = bold_font
                    condition_failed = True
                if (away_ath := ws.cell(row=row, column=col_indices['AWAY_ATH']).value) is not None and away_ath <= -25:
                    ws.cell(row=row, column=col_indices['AWAY_ATH']).fill = no_condition_fill
                    ws.cell(row=row, column=col_indices['AWAY_ATH']).font = bold_font
                    condition_failed = True
                if (roc12M := ws.cell(row=row, column=col_indices['roc12M']).value) is not None and roc12M <= 6.5:
                    ws.cell(row=row, column=col_indices['roc12M']).fill = no_condition_fill
                    ws.cell(row=row, column=col_indices['roc12M']).font = bold_font
                    condition_failed = True
                if (circuit := ws.cell(row=row, column=col_indices['circuit']).value) is not None and circuit >= 20:
                    ws.cell(row=row, column=col_indices['circuit']).fill = no_condition_fill
                    ws.cell(row=row, column=col_indices['circuit']).font = bold_font
                    condition_failed = True
                if (
                roc1M := ws.cell(row=row, column=col_indices['roc1M']).value) is not None and roc12M and roc12M != 0:
                    if roc1M / roc12M * 100 >= 50:
                        ws.cell(row=row, column=col_indices['roc1M']).fill = no_condition_fill
                        ws.cell(row=row, column=col_indices['roc1M']).font = bold_font
                        condition_failed = True
                if close is not None and close <= 30:
                    ws.cell(row=row, column=col_indices['Close']).fill = no_condition_fill
                    ws.cell(row=row, column=col_indices['Close']).font = bold_font
                    condition_failed = True
                if (circuit5 := ws.cell(row=row, column=col_indices['circuit5']).value) is not None and circuit5 > 10:
                    ws.cell(row=row, column=col_indices['circuit5']).fill = no_condition_fill
                    ws.cell(row=row, column=col_indices['circuit5']).font = bold_font
                    condition_failed = True
                if roc12M is not None and roc12M > 1000:
                    ws.cell(row=row, column=col_indices['roc12M']).fill = no_condition_fill
                    ws.cell(row=row, column=col_indices['roc12M']).font = bold_font
                    condition_failed = True
                if condition_failed:
                    ws.cell(row=row, column=col_indices['Ticker']).fill = no_condition_fill

                    # Round off "ATH" column values
                    ath_idx = None
                    for col in range(1, ws.max_column + 1):
                        if ws.cell(row=1, column=col).value == "ATH":  # Search for "ATH" header
                            ath_idx = col
                            break
                    if ath_idx:
                        for row in range(2, ws.max_row + 1):
                            cell = ws.cell(row=row, column=ath_idx)
                            if isinstance(cell.value, (int, float)):
                                cell.value = round(cell.value)

            # Highlight "Rank" column cells <= 75 with light green
            light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_indices['Rank'])
                if cell.value is not None and cell.value <= 75:
                    cell.fill = light_green_fill

            # Save the modified Excel file
            wb.save(file_name)
            print(f"\nExcel file '{file_name}' updated with formatting\n")

        # *********************************************************

        def format_filtered_excel(file_name):
            # Open the written file using openpyxl
            wb = openpyxl.load_workbook(file_name)
            ws = wb["Filtered Stocks"]  # Specify the "Filtered Stocks" sheet

            # Add Borders to All Cells
            thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                                 bottom=Side(style="thin"))

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Freeze the top row
            ws.freeze_panes = 'A2'

            # Format headers
            header_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")  # Dark blue
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Automatically adjust column widths based on content
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2
                ws.column_dimensions[col[0].column_letter].width = adjusted_width

            # Round off "ATH" column values
            ath_idx = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == "ATH":  # Search for "ATH" header
                    ath_idx = col
                    break
            if ath_idx:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=ath_idx)
                    if isinstance(cell.value, (int, float)):
                        cell.value = round(cell.value)

            # Append '%' to "AWAY_ATH" column values
            away_ath_idx = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == "AWAY_ATH":
                    away_ath_idx = col
                    break

            if away_ath_idx:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=away_ath_idx)
                    if isinstance(cell.value, (int, float)):
                        cell.value = f"{cell.value}%"

            # Highlight "Rank" column cells where value <= 75 with light green
            rank_idx = None
            light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == "Rank":
                    rank_idx = col
                    break

            rank_75_count = 0
            if rank_idx:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=rank_idx)
                    if isinstance(cell.value, (int, float)) and cell.value <= 75:
                        cell.fill = light_green_fill
                        rank_75_count += 1

            # Add summary
            total_filtered_stocks = ws.max_row - 1
            ws.append([])  # Empty row
            ws.append(["Summary"])  # Summary heading
            summary_start_row = ws.max_row
            ws.append([f"Total Filtered Stocks:  {total_filtered_stocks}"])
            ws.append([f"Number of Stocks within 75 Rank:  {rank_75_count}"])

            # Apply bold font to the summary
            for row in ws.iter_rows(min_row=summary_start_row, max_row=ws.max_row, min_col=1, max_col=1):
                for cell in row:
                    cell.font = Font(bold=True)

            wb.save(file_name)
            print("\nFiltered Excel file formatted and updated with summary.\n")

        # ********************************************************
        # Format the filename with the lookback date, universe, and other parameters
        excel_file = f"{selected_date.strftime('%Y-%m-%d')}_{U}_{ranking_method}_lookback.xlsx"

        # Save filtered data to Excel
        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            dfStats.to_excel(writer, sheet_name="Unfiltered Stocks", index=True)  # Unfiltered data
            filtered.to_excel(writer, sheet_name="Filtered Stocks", index=True)  # Filtered data

        # Format the Unfiltered Excel file
        format_excel(excel_file)
        # Format the filtered sheet
        format_filtered_excel(excel_file)

        # Download button for the Excel file
        st.download_button(
            label="Download Stock Data as Excel",
            data=open(excel_file, "rb").read(),
            file_name=excel_file,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # **********************************************************

        # Assuming 'dfStats' has 'Rank' as the index and 'final_momentum' filter is applied
        filtered = dfStats[dfStats['final_momentum']].sort_values('Rank', ascending=True)

        # Get the top 75 ranks (using the index for Rank)
        top_75_tickers = filtered[filtered.index <= 75]['Ticker']

        # Fetch the current portfolio from the published CSV (Nifty50 Value)
        portfolio_url = "https://docs.google.com/spreadsheets/d/e/2PACX-1vS4HDgiell4n1kd08OnlzOQobfPzeDtVyWJ8gETFlYbz27qhOmfqKZOoIXZItRQEq5ANATYIcZJm0gk/pub?output=csv"

        # Start the spinner to indicate the process is running
        with st.spinner("Portfolio Rebalancing... Please wait..."):
            # Simulate the delay of fetching and processing data (remove in actual code)

            # Read portfolio data
            portfolio_data = pd.read_csv(portfolio_url)

            # Check if 'Current Portfolio' column exists in the portfolio CSV
            if 'Current Portfolio' not in portfolio_data.columns:
                st.error("Column 'Current Portfolio' not found in the portfolio data.")
            else:
                # Assuming the portfolio has a 'Current Portfolio' column for holdings
                current_portfolio_tickers = portfolio_data['Current Portfolio']

                # Find entry stocks (stocks in top 75 that are not in the current portfolio)
                entry_stocks = top_75_tickers[~top_75_tickers.isin(current_portfolio_tickers)]

                # Find exit stocks (stocks in the current portfolio that are not in the top 75)
                exit_stocks = current_portfolio_tickers[~current_portfolio_tickers.isin(top_75_tickers)]

                # Display results using Streamlit
                st.write("Portfolio Rebalancing:")

                # Limit the number of buy (entry) stocks to match the number of sell (exit) stocks
                num_sells = len(exit_stocks)
                entry_stocks = entry_stocks.head(num_sells)  # Limit Buy tickers to the number of Sell tickers

                # If there are more sells than buys, fill the remaining buys with None or NaN
                if len(entry_stocks) < num_sells:
                    # Use pd.concat to add None values to entry_stocks
                    entry_stocks = pd.concat([entry_stocks, pd.Series([None] * (num_sells - len(entry_stocks)))])

                # Create rebalance table
                rebalance_table = pd.DataFrame({
                    'S.No.': range(1, num_sells + 1),
                    'Sell Stocks': exit_stocks.tolist(),
                    'Buy Stocks': entry_stocks.tolist()
                })

                # Remove rows where both 'Sell' and 'Buy' are None, but keep rows where only one is None
                rebalance_table = rebalance_table[
                    ~((rebalance_table['Sell Stocks'].isna()) & (rebalance_table['Buy Stocks'].isna()))]

                # Set 'S.No.' as the index
                rebalance_table.set_index('S.No.', inplace=True)

                # Display rebalance table
                st.dataframe(rebalance_table)

        # After the spinner ends, show success message
        st.success("Portfolio Rebalancing completed!")


if __name__ == "__main__":
    main()
# ***************************************************************

